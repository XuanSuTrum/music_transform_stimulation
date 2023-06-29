# -*- coding: utf-8 -*-
"""
Created on Wed Jun 28 16:41:54 2023

@author: Administrator
"""

import librosa
import numpy as np
import matplotlib.pyplot as plt
import openpyxl
from matplotlib.cm import get_cmap
import matplotlib.pyplot as plt
# 读取音频文件
audio_file = "d:\Downloads\Jingle Bells-Childrens Christmas Favorites.128.mp3"
y, sr = librosa.load(audio_file)
n_fft=1024
hop_length=512
n_mels=128
# 计算频谱
melspec = librosa.feature.melspectrogram(y=y, sr=sr, S=None,n_fft=1024, hop_length=512, n_mels=128)

#绘画频谱图
logmelspec = librosa.power_to_db(melspec)
plt.figure()
librosa.display.specshow(data=logmelspec, sr=sr, x_axis='time', y_axis='mel')
plt.colorbar(format='%+2.0f dB')
plt.title('Jingle Bells-Childrens Christmas Spectrogram')
plt.show()

# 转换为幅度谱
magnitude = np.abs(melspec)

# 获取每个时间点上能量最强的频率值和相应的能量值
max_freqs = np.argmax(magnitude, axis=0)  # 每列中能量最强的频率的索引
max_magnitude = np.max(magnitude, axis=0)  # 每列中能量最强的频率的能量值

# 根据频率索引获取实际频率值
frequencies = librosa.fft_frequencies(sr=sr, n_fft=1024)

# 存储时间、频率和能量的信息
time_freq_energy = []

# 输出每个时间点上能量最强的频率值和能量值
for t in range(max_freqs.shape[0]):
    max_freq = frequencies[max_freqs[t]]
    max_energy = max_magnitude[t]    
    time = t * hop_length / sr
    time_freq_energy.append((time, max_freq, max_energy))
    print(f"Time: {t * hop_length / sr} s, Frequency: {max_freq} Hz, Energy: {max_energy}")
    
np.savetxt(r'C:\\Users\\Administrator\\Desktop\\code_result\\music_electronics\time_freq_energy.txt', time_freq_energy, fmt='%.6f', delimiter='\t')

# 存储时间-音高数据
time_pitch_data = []

pitch_freq_mapping = {}
workbook = openpyxl.load_workbook('C:\\Users\\Administrator\\Desktop\\code_result\\music_electronics\\pitch_freq_mapping.xlsx')
sheet = workbook.active

for row in sheet.iter_rows(values_only=True):
    pitch = int(row[0])
    min_freq = float(row[1])
    max_freq = float(row[2])
    pitch_freq_mapping[pitch] = (min_freq, max_freq)

# 根据时间、频率和能量数据获取时间-音高数据
for time, freq, energy in time_freq_energy:
    # 寻找音高与频率对应关系
    pitch = None
    for pitch, (min_freq, max_freq) in pitch_freq_mapping.items():
        if min_freq <= freq <= max_freq:
            break
    
    # 将时间、音高和能量添加到时间-音高数据中
    time_pitch_data.append((time, pitch, energy))

# 打印时间-音高数据
for time, pitch, energy in time_pitch_data:
    print(f"Time: {time} s, Pitch: {pitch}, Energy: {energy}")

# 保存时间-音高数据到.xlsx文件
workbook = openpyxl.Workbook()
sheet = workbook.active

for data_row in time_pitch_data:
    sheet.append(data_row)

workbook.save('C:\\Users\\Administrator\\Desktop\\code_result\\music_electronics\\time_pitch_data.xlsx')

merged_data = []  # 合并后的音符数据列表

for time, pitch, energy in time_pitch_data:
    if not merged_data:
        merged_data.append((time, time, pitch, energy))  # 添加第一个音符数据
    else:
        last_start_time, last_end_time, last_pitch, last_energy = merged_data[-1]
        if pitch == last_pitch:
            merged_data[-1] = (last_start_time, time, pitch, last_energy)  # 更新结束时间
        else:
            merged_data.append((time, time, pitch, energy))  # 添加新的音符数据

time_pitch_duration_data = []  # 时间-音高-时长的音符数据列表

for start_time, end_time, pitch, energy in merged_data:
    duration = end_time - start_time  # 计算时长
    time_pitch_duration_data.append((start_time, pitch, duration, energy))
# 打印时间-音高-时长的音符数据
for start_time, pitch, duration, energy in time_pitch_duration_data:
    print(f"Start Time: {start_time} s, Pitch: {pitch}, Duration: {duration} s, Energy: {energy}")

loudness_features = []  # 响度特征数据列表

for start_time, pitch, duration, energy in time_pitch_duration_data:
    if duration != 0:
        # 计算每个频率的能量值按照音的长度取平均值
        loudness = energy / duration
    else:
        # 音长为零，无法计算响度，设置为0或其他合适的值
        loudness = 0

    # 添加时间、音高、音长和响度特征到响度特征数据列表
    loudness_features.append((start_time, pitch, duration, loudness))

# 打印时间、音高、音长和响度特征数据
for start_time, pitch, duration, loudness in loudness_features:
    print(f"Time: {start_time} s, Pitch: {pitch}, Duration: {duration} s, Loudness: {loudness}")
    



